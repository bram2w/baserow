import zipfile
from io import BytesIO
from unittest.mock import MagicMock, patch

from django.test.utils import override_settings

import pytest
from baserow_premium.license.exceptions import FeaturesNotAvailableError
from openpyxl import load_workbook

from baserow.contrib.database.export.handler import ExportHandler
from baserow.contrib.database.export.models import EXPORT_JOB_FINISHED_STATUS
from baserow.contrib.database.rows.handler import RowHandler
from baserow.core.storage import get_default_storage
from baserow.test_utils.helpers import setup_interesting_test_table


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_can_export_every_interesting_different_field_to_json(
    get_storage_mock, premium_data_fixture
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock

    contents = run_export_over_interesting_test_table(
        premium_data_fixture,
        storage_mock,
        {"exporter_type": "json"},
        user_kwargs={"has_active_premium_license": True, "email": "user@example.com"},
    )
    assert (
        contents
        == """[
{
    "id": 1,
    "text": "",
    "long_text": "",
    "url": "",
    "email": "",
    "negative_int": "",
    "positive_int": "",
    "negative_decimal": "",
    "positive_decimal": "",
    "decimal_with_default": "1.8",
    "rating": 0,
    "boolean": false,
    "boolean_with_default": true,
    "datetime_us": "",
    "date_us": "",
    "datetime_eu": "",
    "date_eu": "",
    "datetime_eu_tzone_visible": "",
    "datetime_eu_tzone_hidden": "",
    "last_modified_datetime_us": "01/02/2021 12:00",
    "last_modified_date_us": "01/02/2021",
    "last_modified_datetime_eu": "02/01/2021 12:00",
    "last_modified_date_eu": "02/01/2021",
    "last_modified_datetime_eu_tzone": "02/01/2021 13:00",
    "created_on_datetime_us": "01/02/2021 12:00",
    "created_on_date_us": "01/02/2021",
    "created_on_datetime_eu": "02/01/2021 12:00",
    "created_on_date_eu": "02/01/2021",
    "created_on_datetime_eu_tzone": "02/01/2021 13:00",
    "last_modified_by": "user@example.com",
    "created_by": "user@example.com",
    "duration_hm": "",
    "duration_hms": "",
    "duration_hms_s": "",
    "duration_hms_ss": "",
    "duration_hms_sss": "",
    "duration_dh": "",
    "duration_dhm": "",
    "duration_dhms": "",
    "link_row": [],
    "self_link_row": [],
    "link_row_without_related": [],
    "decimal_link_row": [],
    "file_link_row": [],
    "multiple_collaborators_link_row": [],
    "file": [],
    "single_select": "",
    "single_select_with_default": "BB",
    "multiple_select": [],
    "multiple_select_with_default": [
        "M-1",
        "M-2"
    ],
    "multiple_collaborators": [],
    "phone_number": "",
    "formula_text": "test FORMULA",
    "formula_int": 1,
    "formula_bool": true,
    "formula_decimal": "33.3333333333",
    "formula_dateinterval": "1d 0:00",
    "formula_date": "2020-01-01",
    "formula_singleselect": "",
    "formula_email": "",
    "formula_link_with_label": {
        "url": "https://google.com",
        "label": "label"
    },
    "formula_link_url_only": {
        "url": "https://google.com"
    },
    "formula_multipleselect": [],
    "formula_multiple_collaborators": [],
    "count": 0,
    "rollup": "0.000",
    "duration_rollup_sum": "0:00",
    "duration_rollup_avg": "0:00",
    "lookup": [],
    "multiple_collaborators_lookup": [],
    "uuid": "00000000-0000-4000-8000-000000000001",
    "autonumber": 1,
    "password": "",
    "ai": "",
    "ai_choice": ""
},
{
    "id": 2,
    "text": "text",
    "long_text": "long_text",
    "url": "https://www.google.com",
    "email": "test@example.com",
    "negative_int": -1,
    "positive_int": 1,
    "negative_decimal": "-1.2",
    "positive_decimal": "1.2",
    "decimal_with_default": "1.8",
    "rating": 3,
    "boolean": true,
    "boolean_with_default": true,
    "datetime_us": "02/01/2020 01:23",
    "date_us": "02/01/2020",
    "datetime_eu": "01/02/2020 01:23",
    "date_eu": "01/02/2020",
    "datetime_eu_tzone_visible": "01/02/2020 02:23",
    "datetime_eu_tzone_hidden": "01/02/2020 02:23",
    "last_modified_datetime_us": "01/02/2021 12:00",
    "last_modified_date_us": "01/02/2021",
    "last_modified_datetime_eu": "02/01/2021 12:00",
    "last_modified_date_eu": "02/01/2021",
    "last_modified_datetime_eu_tzone": "02/01/2021 13:00",
    "created_on_datetime_us": "01/02/2021 12:00",
    "created_on_date_us": "01/02/2021",
    "created_on_datetime_eu": "02/01/2021 12:00",
    "created_on_date_eu": "02/01/2021",
    "created_on_datetime_eu_tzone": "02/01/2021 13:00",
    "last_modified_by": "user@example.com",
    "created_by": "user@example.com",
    "duration_hm": "1:01",
    "duration_hms": "1:01:06",
    "duration_hms_s": "1:01:06.6",
    "duration_hms_ss": "1:01:06.66",
    "duration_hms_sss": "1:01:06.666",
    "duration_dh": "1d 1h",
    "duration_dhm": "1d 1:01",
    "duration_dhms": "1d 1:01:06",
    "link_row": [
        "linked_row_1",
        "linked_row_2",
        ""
    ],
    "self_link_row": [
        "unnamed row 1"
    ],
    "link_row_without_related": [
        "linked_row_1",
        "linked_row_2"
    ],
    "decimal_link_row": [
        "1.234",
        "-123.456",
        "unnamed row 3"
    ],
    "file_link_row": [
        [
            {
                "visible_name": "name.txt",
                "url": "http://localhost:8000/media/user_files/test_hash.txt"
            }
        ],
        "unnamed row 2"
    ],
    "multiple_collaborators_link_row": [
        [
            "User2 <user2@example.com>",
            "User3 <user3@example.com>"
        ],
        [
            "User2 <user2@example.com>"
        ]
    ],
    "file": [
        {
            "visible_name": "a.txt",
            "url": "http://localhost:8000/media/user_files/hashed_name.txt"
        },
        {
            "visible_name": "b.txt",
            "url": "http://localhost:8000/media/user_files/other_name.txt"
        }
    ],
    "single_select": "A",
    "single_select_with_default": "BB",
    "multiple_select": [
        "D",
        "C",
        "E"
    ],
    "multiple_select_with_default": [
        "M-1",
        "M-2"
    ],
    "multiple_collaborators": [
        "User2 <user2@example.com>",
        "User3 <user3@example.com>"
    ],
    "phone_number": "+4412345678",
    "formula_text": "test FORMULA",
    "formula_int": 1,
    "formula_bool": true,
    "formula_decimal": "33.3333333333",
    "formula_dateinterval": "1d 0:00",
    "formula_date": "2020-01-01",
    "formula_singleselect": "A",
    "formula_email": "test@example.com",
    "formula_link_with_label": {
        "url": "https://google.com",
        "label": "label"
    },
    "formula_link_url_only": {
        "url": "https://google.com"
    },
    "formula_multipleselect": [
        "C",
        "D",
        "E"
    ],
    "formula_multiple_collaborators": [
        "User2 <user2@example.com>",
        "User3 <user3@example.com>"
    ],
    "count": 3,
    "rollup": "-122.222",
    "duration_rollup_sum": "0:04",
    "duration_rollup_avg": "0:02",
    "lookup": [
        "linked_row_1",
        "linked_row_2",
        ""
    ],
    "multiple_collaborators_lookup": [
        [
            "User2 <user2@example.com>",
            "User3 <user3@example.com>"
        ],
        [
            "User2 <user2@example.com>"
        ]
    ],
    "uuid": "00000000-0000-4000-8000-000000000002",
    "autonumber": 2,
    "password": true,
    "ai": "I'm an AI.",
    "ai_choice": "Object"
}
]
"""
    )


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_cannot_export_json_without_premium_license(
    get_storage_mock, premium_data_fixture
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    with pytest.raises(FeaturesNotAvailableError):
        run_export_over_interesting_test_table(
            premium_data_fixture, storage_mock, {"exporter_type": "json"}
        )


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_cannot_export_json_without_premium_license_for_group(
    get_storage_mock, premium_data_fixture, alternative_per_workspace_license_service
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    # Setting the group id to `0` will make sure that the user doesn't have
    # premium access to the group.
    user = premium_data_fixture.create_user(has_active_premium_license=True)
    alternative_per_workspace_license_service.restrict_user_premium_to(user, [0])
    with pytest.raises(FeaturesNotAvailableError):
        run_export_over_interesting_test_table(
            premium_data_fixture, storage_mock, {"exporter_type": "json"}, user=user
        )


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_if_duplicate_field_names_json_export(get_storage_mock, premium_data_fixture):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    user = premium_data_fixture.create_user(has_active_premium_license=True)
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)
    premium_data_fixture.create_text_field(table=table, name="name", order=1)
    premium_data_fixture.create_text_field(table=table, name="name", order=2)
    premium_data_fixture.create_text_field(table=table, name="name", order=3)
    premium_data_fixture.create_text_field(table=table, name='Another"name', order=4)
    premium_data_fixture.create_text_field(table=table, name='Another"name', order=5)
    row_handler = RowHandler()
    row_handler.create_row(user=user, table=table)
    job, contents = run_export_job_with_mock_storage(
        table, None, storage_mock, user, {"exporter_type": "json"}
    )
    assert (
        contents
        == """[
{
    "id": 1,
    "name": "",
    "name 2": "",
    "name 3": "",
    "Another\\"name": "",
    "Another\\"name 2": ""
}
]
"""
    )


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_can_export_every_interesting_different_field_to_xml(
    get_storage_mock, premium_data_fixture
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    xml = run_export_over_interesting_test_table(
        premium_data_fixture,
        storage_mock,
        {"exporter_type": "xml"},
        user_kwargs={"has_active_premium_license": True, "email": "user@example.com"},
    )
    expected_xml = """<?xml version="1.0" encoding="utf-8" ?>
<rows>
   <row>
      <id>1</id>
      <text/>
      <long-text/>
      <url/>
      <email/>
      <negative-int/>
      <positive-int/>
      <negative-decimal/>
      <positive-decimal/>
      <decimal-with-default>1.8</decimal-with-default>
      <rating>0</rating>
      <boolean>false</boolean>
      <boolean-with-default>true</boolean-with-default>
      <datetime-us/>
      <date-us/>
      <datetime-eu/>
      <date-eu/>
      <datetime-eu-tzone-visible/>
      <datetime-eu-tzone-hidden/>
      <last-modified-datetime-us>01/02/2021 12:00</last-modified-datetime-us>
      <last-modified-date-us>01/02/2021</last-modified-date-us>
      <last-modified-datetime-eu>02/01/2021 12:00</last-modified-datetime-eu>
      <last-modified-date-eu>02/01/2021</last-modified-date-eu>
      <last-modified-datetime-eu-tzone>02/01/2021 13:00</last-modified-datetime-eu-tzone>
      <created-on-datetime-us>01/02/2021 12:00</created-on-datetime-us>
      <created-on-date-us>01/02/2021</created-on-date-us>
      <created-on-datetime-eu>02/01/2021 12:00</created-on-datetime-eu>
      <created-on-date-eu>02/01/2021</created-on-date-eu>
      <created-on-datetime-eu-tzone>02/01/2021 13:00</created-on-datetime-eu-tzone>
      <last-modified-by>user@example.com</last-modified-by>
      <created-by>user@example.com</created-by>
      <duration-hm/>
      <duration-hms/>
      <duration-hms-s/>
      <duration-hms-ss/>
      <duration-hms-sss/>
      <duration-dh/>
      <duration-dhm/>
      <duration-dhms/>
      <link-row/>
      <self-link-row/>
      <link-row-without-related/>
      <decimal-link-row/>
      <file-link-row/>
      <multiple-collaborators-link-row/>
      <file/>
      <single-select/>
      <single-select-with-default>BB</single-select-with-default>
      <multiple-select/>
      <multiple-select-with-default>
         <item>M-1</item>
         <item>M-2</item>
      </multiple-select-with-default>
      <multiple-collaborators/>
      <phone-number/>
      <formula-text>test FORMULA</formula-text>
      <formula-int>1</formula-int>
      <formula-bool>true</formula-bool>
      <formula-decimal>33.3333333333</formula-decimal>
      <formula-dateinterval>1d 0:00</formula-dateinterval>
      <formula-date>2020-01-01</formula-date>
      <formula-singleselect/>
      <formula-email/>
      <formula-link-with-label>
         <url>https://google.com</url>
         <label>label</label>
      </formula-link-with-label>
      <formula-link-url-only>
         <url>https://google.com</url>
      </formula-link-url-only>
      <formula-multipleselect/>
      <formula-multiple-collaborators/>
      <count>0</count>
      <rollup>0.000</rollup>
      <duration-rollup-sum>0:00</duration-rollup-sum>
      <duration-rollup-avg>0:00</duration-rollup-avg>
      <lookup/>
      <multiple-collaborators-lookup/>
      <uuid>00000000-0000-4000-8000-000000000001</uuid>
      <autonumber>1</autonumber>
      <password/>
      <ai/>
      <ai-choice/>
   </row>
   <row>
      <id>2</id>
      <text>text</text>
      <long-text>long_text</long-text>
      <url>https://www.google.com</url>
      <email>test@example.com</email>
      <negative-int>-1</negative-int>
      <positive-int>1</positive-int>
      <negative-decimal>-1.2</negative-decimal>
      <positive-decimal>1.2</positive-decimal>
      <decimal-with-default>1.8</decimal-with-default>
      <rating>3</rating>
      <boolean>true</boolean>
      <boolean-with-default>true</boolean-with-default>
      <datetime-us>02/01/2020 01:23</datetime-us>
      <date-us>02/01/2020</date-us>
      <datetime-eu>01/02/2020 01:23</datetime-eu>
      <date-eu>01/02/2020</date-eu>
      <datetime-eu-tzone-visible>01/02/2020 02:23</datetime-eu-tzone-visible>
      <datetime-eu-tzone-hidden>01/02/2020 02:23</datetime-eu-tzone-hidden>
      <last-modified-datetime-us>01/02/2021 12:00</last-modified-datetime-us>
      <last-modified-date-us>01/02/2021</last-modified-date-us>
      <last-modified-datetime-eu>02/01/2021 12:00</last-modified-datetime-eu>
      <last-modified-date-eu>02/01/2021</last-modified-date-eu>
      <last-modified-datetime-eu-tzone>02/01/2021 13:00</last-modified-datetime-eu-tzone>
      <created-on-datetime-us>01/02/2021 12:00</created-on-datetime-us>
      <created-on-date-us>01/02/2021</created-on-date-us>
      <created-on-datetime-eu>02/01/2021 12:00</created-on-datetime-eu>
      <created-on-date-eu>02/01/2021</created-on-date-eu>
      <created-on-datetime-eu-tzone>02/01/2021 13:00</created-on-datetime-eu-tzone>
      <last-modified-by>user@example.com</last-modified-by>
      <created-by>user@example.com</created-by>
      <duration-hm>1:01</duration-hm>
      <duration-hms>1:01:06</duration-hms>
      <duration-hms-s>1:01:06.6</duration-hms-s>
      <duration-hms-ss>1:01:06.66</duration-hms-ss>
      <duration-hms-sss>1:01:06.666</duration-hms-sss>
      <duration-dh>1d 1h</duration-dh>
      <duration-dhm>1d 1:01</duration-dhm>
      <duration-dhms>1d 1:01:06</duration-dhms>
      <link-row>
         <item>linked_row_1</item>
         <item>linked_row_2</item>
         <item/>
      </link-row>
      <self-link-row>
         <item>unnamed row 1</item>
      </self-link-row>
      <link-row-without-related>
         <item>linked_row_1</item>
         <item>linked_row_2</item>
      </link-row-without-related>
      <decimal-link-row>
         <item>1.234</item>
         <item>-123.456</item>
         <item>unnamed row 3</item>
      </decimal-link-row>
      <file-link-row>
         <item>
            <item>
               <visible_name>name.txt</visible_name>
               <url>http://localhost:8000/media/user_files/test_hash.txt</url>
            </item>
         </item>
         <item>unnamed row 2</item>
      </file-link-row>
      <multiple-collaborators-link-row>
        <item>
            <item>User2 &lt;user2@example.com&gt;</item>
            <item>User3 &lt;user3@example.com&gt;</item>
        </item>
        <item>
            <item>User2 &lt;user2@example.com&gt;</item>
        </item>
      </multiple-collaborators-link-row>
      <file>
         <item>
            <visible_name>a.txt</visible_name>
            <url>http://localhost:8000/media/user_files/hashed_name.txt</url>
         </item>
         <item>
            <visible_name>b.txt</visible_name>
            <url>http://localhost:8000/media/user_files/other_name.txt</url>
         </item>
      </file>
      <single-select>A</single-select>
      <single-select-with-default>BB</single-select-with-default>
      <multiple-select>
         <item>D</item>
         <item>C</item>
         <item>E</item>
      </multiple-select>
      <multiple-select-with-default>
         <item>M-1</item>
         <item>M-2</item>
      </multiple-select-with-default>
      <multiple-collaborators>
         <item>User2 &lt;user2@example.com&gt;</item>
         <item>User3 &lt;user3@example.com&gt;</item>
      </multiple-collaborators>
      <phone-number>+4412345678</phone-number>
      <formula-text>test FORMULA</formula-text>
      <formula-int>1</formula-int>
      <formula-bool>true</formula-bool>
      <formula-decimal>33.3333333333</formula-decimal>
      <formula-dateinterval>1d 0:00</formula-dateinterval>
      <formula-date>2020-01-01</formula-date>
      <formula-singleselect>A</formula-singleselect>
      <formula-email>test@example.com</formula-email>
      <formula-link-with-label>
         <url>https://google.com</url>
         <label>label</label>
      </formula-link-with-label>
      <formula-link-url-only>
         <url>https://google.com</url>
      </formula-link-url-only>
      <formula-multipleselect>
         <item>C</item>
         <item>D</item>
         <item>E</item>
      </formula-multipleselect>
      <formula-multiple-collaborators>
         <item>User2 &lt;user2@example.com&gt;</item>
         <item>User3 &lt;user3@example.com&gt;</item>
      </formula-multiple-collaborators>
      <count>3</count>
      <rollup>-122.222</rollup>
      <duration-rollup-sum>0:04</duration-rollup-sum>
      <duration-rollup-avg>0:02</duration-rollup-avg>
      <lookup>
         <item>linked_row_1</item>
         <item>linked_row_2</item>
         <item/>
      </lookup>
      <multiple-collaborators-lookup>
        <item>
            <item>User2 &lt;user2@example.com&gt;</item>
            <item>User3 &lt;user3@example.com&gt;</item>
        </item>
        <item>
            <item>User2 &lt;user2@example.com&gt;</item>
        </item>
      </multiple-collaborators-lookup>
      <uuid>00000000-0000-4000-8000-000000000002</uuid>
      <autonumber>2</autonumber>
      <password>true</password>
      <ai>I'm an AI.</ai>
      <ai-choice>Object</ai-choice>
   </row>
</rows>
"""
    assert strip_indents_and_newlines(xml) == strip_indents_and_newlines(expected_xml)


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_if_xml_duplicate_name_and_value_are_escaped(
    get_storage_mock, premium_data_fixture
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    user = premium_data_fixture.create_user(has_active_premium_license=True)
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)
    text = premium_data_fixture.create_text_field(table=table, name="<name>", order=0)
    premium_data_fixture.create_text_field(table=table, name="name", order=1)
    premium_data_fixture.create_text_field(table=table, name="Another name", order=2)
    premium_data_fixture.create_text_field(table=table, name="Another@name", order=3)
    empty_1 = premium_data_fixture.create_text_field(table=table, name="@", order=4)
    empty_2 = premium_data_fixture.create_text_field(table=table, name="", order=5)
    premium_data_fixture.create_text_field(table=table, name="1", order=6)
    row_handler = RowHandler()
    row_handler.create_row(
        user=user,
        table=table,
        values={f"field_{text.id}": "<value>"},
    )
    job, contents = run_export_job_with_mock_storage(
        table, None, storage_mock, user, {"exporter_type": "xml"}
    )
    assert strip_indents_and_newlines(contents) == strip_indents_and_newlines(
        f"""
<?xml version="1.0" encoding="utf-8" ?>
<rows>
  <row>
    <id>1</id>
    <name>&lt;value&gt;</name>
    <name-2/>
    <Another-name/>
    <Another-name-2/>
    <field-{empty_1.id}/>
    <field-{empty_2.id}/>
    <field-1/>
  </row>
</rows>
"""
    )


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_cannot_export_xml_without_premium_license(
    get_storage_mock, premium_data_fixture
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    with pytest.raises(FeaturesNotAvailableError):
        run_export_over_interesting_test_table(
            premium_data_fixture, storage_mock, {"exporter_type": "xml"}
        )


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_cannot_export_xml_without_premium_license_for_group(
    get_storage_mock, premium_data_fixture, alternative_per_workspace_license_service
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    # Setting the group id to `0` will make sure that the user doesn't have
    # premium access to the group.
    user = premium_data_fixture.create_user(has_active_premium_license=True)
    alternative_per_workspace_license_service.restrict_user_premium_to(user, [0])
    with pytest.raises(FeaturesNotAvailableError):
        run_export_over_interesting_test_table(
            premium_data_fixture, storage_mock, {"exporter_type": "xml"}, user=user
        )


def strip_indents_and_newlines(xml):
    return "".join([line.strip() for line in xml.split("\n")])


def run_export_over_interesting_test_table(
    premium_data_fixture, storage_mock, options, user_kwargs=None, user=None
):
    table, user, _, _, context = setup_interesting_test_table(
        premium_data_fixture, user_kwargs=user_kwargs, user=user
    )
    grid_view = premium_data_fixture.create_grid_view(table=table)
    job, contents = run_export_job_with_mock_storage(
        table, grid_view, storage_mock, user, options
    )
    return contents


def run_export_job_with_mock_storage(
    table, grid_view, storage_mock, user, options=None
):
    if options is None:
        options = {"exporter_type": "csv"}

    if "export_charset" not in options:
        options["export_charset"] = "utf-8"

    stub_file = BytesIO()
    storage_mock.open.return_value = stub_file
    close = stub_file.close
    stub_file.close = lambda: None
    handler = ExportHandler()
    job = handler.create_pending_export_job(user, table, grid_view, options)
    handler.run_export_job(job)
    actual = stub_file.getvalue()
    if options["export_charset"]:
        actual = actual.decode(options["export_charset"])
    close()
    return job, actual


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_cannot_export_excel_without_premium_license_for_group(
    get_storage_mock, premium_data_fixture, alternative_per_workspace_license_service
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock
    # Setting the group id to `0` will make sure that the user doesn't have
    # premium access to the group.
    user = premium_data_fixture.create_user(has_active_premium_license=True)
    alternative_per_workspace_license_service.restrict_user_premium_to(user, [0])
    with pytest.raises(FeaturesNotAvailableError):
        run_export_over_interesting_test_table(
            premium_data_fixture, storage_mock, {"exporter_type": "excel"}, user=user
        )


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_can_export_every_interesting_different_field_to_excel(
    get_storage_mock, premium_data_fixture
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock

    contents = run_export_over_interesting_test_table(
        premium_data_fixture,
        storage_mock,
        {
            "exporter_type": "excel",
            "export_charset": None,
            "excel_include_header": True,
        },
        user_kwargs={"has_active_premium_license": True, "email": "user@example.com"},
    )

    excel_file = BytesIO(contents)
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    # Define the expected headers and first two rows of data
    expected_headers = [
        "id",
        "text",
        "long_text",
        "url",
        "email",
        "negative_int",
        "positive_int",
        "negative_decimal",
        "positive_decimal",
        "decimal_with_default",
        "rating",
        "boolean",
        "boolean_with_default",
        "datetime_us",
        "date_us",
        "datetime_eu",
        "date_eu",
        "datetime_eu_tzone_visible",
        "datetime_eu_tzone_hidden",
        "last_modified_datetime_us",
        "last_modified_date_us",
        "last_modified_datetime_eu",
        "last_modified_date_eu",
        "last_modified_datetime_eu_tzone",
        "created_on_datetime_us",
        "created_on_date_us",
        "created_on_datetime_eu",
        "created_on_date_eu",
        "created_on_datetime_eu_tzone",
        "last_modified_by",
        "created_by",
        "duration_hm",
        "duration_hms",
        "duration_hms_s",
        "duration_hms_ss",
        "duration_hms_sss",
        "duration_dh",
        "duration_dhm",
        "duration_dhms",
        "link_row",
        "self_link_row",
        "link_row_without_related",
        "decimal_link_row",
        "file_link_row",
        "multiple_collaborators_link_row",
        "file",
        "single_select",
        "single_select_with_default",
        "multiple_select",
        "multiple_select_with_default",
        "multiple_collaborators",
        "phone_number",
        "formula_text",
        "formula_int",
        "formula_bool",
        "formula_decimal",
        "formula_dateinterval",
        "formula_date",
        "formula_singleselect",
        "formula_email",
        "formula_link_with_label",
        "formula_link_url_only",
        "formula_multipleselect",
        "formula_multiple_collaborators",
        "count",
        "rollup",
        "duration_rollup_sum",
        "duration_rollup_avg",
        "lookup",
        "multiple_collaborators_lookup",
        "uuid",
        "autonumber",
        "password",
        "ai",
        "ai_choice",
    ]

    expected_first_row = [
        "1",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "1.8",
        "0",
        "False",
        "True",
        None,
        None,
        None,
        None,
        None,
        None,
        "01/02/2021 12:00",
        "01/02/2021",
        "02/01/2021 12:00",
        "02/01/2021",
        "02/01/2021 13:00",
        "01/02/2021 12:00",
        "01/02/2021",
        "02/01/2021 12:00",
        "02/01/2021",
        "02/01/2021 13:00",
        "user@example.com",
        "user@example.com",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "BB",
        None,
        "M-1,M-2",
        None,
        None,
        "test FORMULA",
        "1",
        "True",
        "33.3333333333",
        "1d 0:00",
        "2020-01-01",
        None,
        None,
        "label (https://google.com)",
        "https://google.com",
        None,
        None,
        "0",
        "0.000",
        "0:00",
        "0:00",
        None,
        None,
        "00000000-0000-4000-8000-000000000001",
        "1",
        None,
        None,
        None,
    ]

    expected_second_row = [
        "2",
        "text",
        "long_text",
        "https://www.google.com",
        "test@example.com",
        "-1",
        "1",
        "-1.2",
        "1.2",
        "1.8",
        "3",
        "True",
        "True",
        "02/01/2020 01:23",
        "02/01/2020",
        "01/02/2020 01:23",
        "01/02/2020",
        "01/02/2020 02:23",
        "01/02/2020 02:23",
        "01/02/2021 12:00",
        "01/02/2021",
        "02/01/2021 12:00",
        "02/01/2021",
        "02/01/2021 13:00",
        "01/02/2021 12:00",
        "01/02/2021",
        "02/01/2021 12:00",
        "02/01/2021",
        "02/01/2021 13:00",
        "user@example.com",
        "user@example.com",
        "1:01",
        "1:01:06",
        "1:01:06.6",
        "1:01:06.66",
        "1:01:06.666",
        "1d 1h",
        "1d 1:01",
        "1d 1:01:06",
        "linked_row_1,linked_row_2,",
        "unnamed row 1",
        "linked_row_1,linked_row_2",
        "1.234,-123.456,unnamed row 3",
        "name.txt (http://localhost:8000/media/user_files/test_hash.txt),unnamed row 2",
        '"User2 <user2@example.com>,User3 <user3@example.com>",User2 <user2@example.com>',
        "a.txt (http://localhost:8000/media/user_files/hashed_name.txt),b.txt (http://localhost:8000/media/user_files/other_name.txt)",
        "A",
        "BB",
        "D,C,E",
        "M-1,M-2",
        "User2 <user2@example.com>,User3 <user3@example.com>",
        "+4412345678",
        "test FORMULA",
        "1",
        "True",
        "33.3333333333",
        "1d 0:00",
        "2020-01-01",
        "A",
        "test@example.com",
        "label (https://google.com)",
        "https://google.com",
        "C,D,E",
        "User2 <user2@example.com>,User3 <user3@example.com>",
        "3",
        "-122.222",
        "0:04",
        "0:02",
        "linked_row_1,linked_row_2,",
        '"User2 <user2@example.com>,User3 <user3@example.com>",User2 <user2@example.com>',
        "00000000-0000-4000-8000-000000000002",
        "2",
        "True",
        "I'm an AI.",
        "Object",
    ]

    # Verify headers
    actual_headers = [cell.value for cell in worksheet[1]]
    assert actual_headers == expected_headers

    # Verify first row
    actual_first_row = [cell.value for cell in worksheet[2]]
    assert actual_first_row == expected_first_row

    # Verify second row
    actual_second_row = [cell.value for cell in worksheet[3]]
    assert actual_second_row == expected_second_row


@pytest.mark.django_db
@override_settings(DEBUG=True)
@patch("baserow.core.storage.get_default_storage")
def test_can_export_every_interesting_different_field_to_excel_without_header(
    get_storage_mock, premium_data_fixture
):
    storage_mock = MagicMock()
    get_storage_mock.return_value = storage_mock

    contents = run_export_over_interesting_test_table(
        premium_data_fixture,
        storage_mock,
        {"exporter_type": "excel", "export_charset": None},
        user_kwargs={"has_active_premium_license": True, "email": "user@example.com"},
    )

    excel_file = BytesIO(contents)
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    # Verify headers
    actual_headers = [cell.value for cell in worksheet[1]]
    assert actual_headers[0] == "1"


@pytest.mark.django_db
@override_settings(DEBUG=True)
def test_can_trigger_export_files_with_premium_license(premium_data_fixture):
    options = {"exporter_type": "file"}

    user = premium_data_fixture.create_user(has_active_premium_license=True)
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)
    grid_view = premium_data_fixture.create_grid_view(table=table)

    ExportHandler().create_pending_export_job(user, table, grid_view, options)


@pytest.mark.django_db
@override_settings(DEBUG=True)
def test_cannot_trigger_export_files_without_premium_license(premium_data_fixture):
    options = {"exporter_type": "file"}

    user = premium_data_fixture.create_user()
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)
    grid_view = premium_data_fixture.create_grid_view(table=table)

    with pytest.raises(FeaturesNotAvailableError):
        ExportHandler().create_pending_export_job(user, table, grid_view, options)


@pytest.mark.django_db
@override_settings(DEBUG=True)
def test_cannnot_trigger_export_files_without_premium_license_for_group(
    premium_data_fixture, alternative_per_workspace_license_service
):
    options = {"exporter_type": "file"}
    user = premium_data_fixture.create_user(has_active_premium_license=True)
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)

    alternative_per_workspace_license_service.restrict_user_premium_to(user, [0])

    grid_view = premium_data_fixture.create_grid_view(table=table)

    with pytest.raises(FeaturesNotAvailableError):
        ExportHandler().create_pending_export_job(user, table, grid_view, options)


@pytest.mark.django_db
@override_settings(DEBUG=True)
def test_export_empty_files_does_not_fail(premium_data_fixture):
    storage = get_default_storage()

    options = {"exporter_type": "file"}
    user = premium_data_fixture.create_user(has_active_premium_license=True)
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)
    grid_view = premium_data_fixture.create_grid_view(table=table)

    handler = ExportHandler()
    job = handler.create_pending_export_job(user, table, grid_view, options)

    handler.run_export_job(job)

    job.refresh_from_db()

    assert job.state == EXPORT_JOB_FINISHED_STATUS
    assert job.exported_file_name is not None

    file_name = job.exported_file_name

    export_file_path = ExportHandler.export_file_path(file_name)

    assert storage.exists(export_file_path)

    with storage.open(export_file_path, "rb") as zip_file:
        with zipfile.ZipFile(zip_file) as archive:
            file_list = archive.namelist()
            assert len(file_list) == 0


@pytest.mark.django_db
@override_settings(DEBUG=True)
def test_export_non_empty_files_flat(premium_data_fixture, use_tmp_media_root):
    storage = get_default_storage()

    user = premium_data_fixture.create_user(has_active_premium_license=True)
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)
    file_field = premium_data_fixture.create_file_field(table=table, name="File")

    row_handler = RowHandler()
    model = table.get_model()

    user_file_1 = premium_data_fixture.create_user_file(
        original_name="a.txt",
        size=16,
        mime_type="text/plain",
        uploaded_by=user,
    )
    premium_data_fixture.save_content_in_user_file(
        user_file_1, storage, "Content of a.txt"
    )

    user_file_2 = premium_data_fixture.create_user_file(
        original_name="b.txt",
        size=16,
        mime_type="text/plain",
        uploaded_by=user,
    )
    premium_data_fixture.save_content_in_user_file(
        user_file_2, storage, "Content of b.txt"
    )

    user_file_3 = premium_data_fixture.create_user_file(
        original_name="c.txt",
        size=16,
        mime_type="text/plain",
        uploaded_by=user,
    )
    premium_data_fixture.save_content_in_user_file(
        user_file_3, storage, "Content of c.txt"
    )

    row_handler.create_row(
        user=user,
        table=table,
        values={
            f"field_{file_field.id}": [
                {"name": user_file_1.name, "visible_name": "a.txt"},
                {"name": user_file_2.name, "visible_name": "b.txt"},
            ]
        },
        model=model,
    )
    row_handler.create_row(
        user=user,
        table=table,
        values={
            f"field_{file_field.id}": [
                {"name": user_file_2.name, "visible_name": "b.txt"},
                {"name": user_file_3.name, "visible_name": "c.txt"},
            ]
        },
        model=model,
    )

    grid_view = premium_data_fixture.create_grid_view(table=table)

    options = {"exporter_type": "file", "organize_files": False}
    handler = ExportHandler()
    job = handler.create_pending_export_job(user, table, grid_view, options)

    handler.run_export_job(job)

    job.refresh_from_db()

    assert job.state == EXPORT_JOB_FINISHED_STATUS
    assert job.exported_file_name is not None

    file_name = job.exported_file_name

    export_file_path = ExportHandler.export_file_path(file_name)

    assert storage.exists(export_file_path)

    with storage.open(export_file_path, "rb") as zip_file:
        with zipfile.ZipFile(zip_file) as archive:
            file_list = archive.namelist()
            assert len(file_list) == 3
            assert user_file_1.name in file_list
            assert user_file_2.name in file_list
            assert user_file_3.name in file_list


@pytest.mark.django_db
@override_settings(DEBUG=True)
def test_export_non_empty_files_group_by_row(premium_data_fixture, use_tmp_media_root):
    storage = get_default_storage()

    user = premium_data_fixture.create_user(has_active_premium_license=True)
    database = premium_data_fixture.create_database_application(user=user)
    table = premium_data_fixture.create_database_table(database=database)
    file_field = premium_data_fixture.create_file_field(table=table, name="File")

    row_handler = RowHandler()
    model = table.get_model()

    user_file_1 = premium_data_fixture.create_user_file(
        original_name="a.txt",
        size=16,
        mime_type="text/plain",
        uploaded_by=user,
    )
    premium_data_fixture.save_content_in_user_file(
        user_file_1, storage, "Content of a.txt"
    )

    user_file_2 = premium_data_fixture.create_user_file(
        original_name="b.txt",
        size=16,
        mime_type="text/plain",
        uploaded_by=user,
    )
    premium_data_fixture.save_content_in_user_file(
        user_file_2, storage, "Content of b.txt"
    )

    user_file_3 = premium_data_fixture.create_user_file(
        original_name="c.txt",
        size=16,
        mime_type="text/plain",
        uploaded_by=user,
    )
    premium_data_fixture.save_content_in_user_file(
        user_file_3, storage, "Content of c.txt"
    )

    row_1 = row_handler.create_row(
        user=user,
        table=table,
        values={
            f"field_{file_field.id}": [
                {"name": user_file_1.name, "visible_name": "a.txt"},
                {"name": user_file_2.name, "visible_name": "b.txt"},
            ]
        },
        model=model,
    )
    row_2 = row_handler.create_row(
        user=user,
        table=table,
        values={
            f"field_{file_field.id}": [
                {"name": user_file_2.name, "visible_name": "b.txt"},
                {"name": user_file_3.name, "visible_name": "c.txt"},
            ]
        },
        model=model,
    )

    grid_view = premium_data_fixture.create_grid_view(table=table)

    options = {"exporter_type": "file", "organize_files": True}
    handler = ExportHandler()
    job = handler.create_pending_export_job(user, table, grid_view, options)

    handler.run_export_job(job)

    job.refresh_from_db()

    assert job.state == EXPORT_JOB_FINISHED_STATUS
    assert job.exported_file_name is not None

    file_name = job.exported_file_name

    export_file_path = ExportHandler.export_file_path(file_name)

    assert storage.exists(export_file_path)

    with storage.open(export_file_path, "rb") as zip_file:
        with zipfile.ZipFile(zip_file) as archive:
            file_list = archive.namelist()
            assert len(file_list) == 4
            assert f"row_{row_1.id}/{user_file_1.name}" in file_list
            assert f"row_{row_1.id}/{user_file_2.name}" in file_list
            assert f"row_{row_2.id}/{user_file_2.name}" in file_list
            assert f"row_{row_2.id}/{user_file_3.name}" in file_list
